"""
Generador de informes de auditoría y reportes ejecutivos en Excel (.xlsx) para el Módulo de Opacidad.
Diseñado para fiscalizaciones de la autoridad marítima/ambiental y control de gestión de mantenimiento STI.
"""
from datetime import datetime
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from django.utils import timezone
from equipos.models import Equipo, TipoEquipo
from opacidad.models import MedicionOpacidad, ProyeccionOpacidad
from core.models import ConfiguracionSistema


def generar_reporte_excel_opacidad(ano=None, periodo=None, tipo_codigo=None, usuario=None):
    """
    Genera un libro Excel (.xlsx) profesional con 3 hojas:
      1. Resumen Ejecutivo (KPIs, Cobertura, Semáforo de Cumplimiento por Flota y Periodo)
      2. Histórico de Mediciones (Detalle completo de cada test realizado)
      3. Estado Actual de la Flota (Último control y estado predictivo de cada equipo)
    Devuelve un BytesIO con el archivo listo para enviar por HTTP.
    """
    wb = openpyxl.Workbook()
    
    # ----------------------------------------------------
    # ESTILOS CORPORATIVOS STI / TIREWATCH
    # ----------------------------------------------------
    FONT_FAMILY = "Segoe UI"
    
    # Fuentes
    f_title = Font(name=FONT_FAMILY, size=16, bold=True, color="0F172A")
    f_subtitle = Font(name=FONT_FAMILY, size=10, italic=True, color="475569")
    f_section = Font(name=FONT_FAMILY, size=12, bold=True, color="0F172A")
    f_header = Font(name=FONT_FAMILY, size=10, bold=True, color="FFFFFF")
    f_kpi_num = Font(name=FONT_FAMILY, size=18, bold=True, color="0F172A")
    f_kpi_lbl = Font(name=FONT_FAMILY, size=9, bold=True, color="64748B")
    f_bold = Font(name=FONT_FAMILY, size=9, bold=True, color="0F172A")
    f_normal = Font(name=FONT_FAMILY, size=9, color="1E293B")
    f_code = Font(name="Consolas", size=9, color="0F172A")
    
    # Fills
    fill_header_navy = PatternFill("solid", fgColor="1E293B")
    fill_header_cyan = PatternFill("solid", fgColor="0284C7")
    fill_kpi_card = PatternFill("solid", fgColor="F8FAFC")
    fill_zebra = PatternFill("solid", fgColor="F1F5F9")
    fill_green = PatternFill("solid", fgColor="DCFCE7")  # Aprobado / OK
    fill_yellow = PatternFill("solid", fgColor="FEF3C7") # Atención
    fill_red = PatternFill("solid", fgColor="FEE2E2")    # Crítico / Reprobado
    
    # Text colors for status
    f_green = Font(name=FONT_FAMILY, size=9, bold=True, color="166534")
    f_yellow = Font(name=FONT_FAMILY, size=9, bold=True, color="92400E")
    f_red = Font(name=FONT_FAMILY, size=9, bold=True, color="991B1B")

    # Bordes
    thin_line = Side(style="thin", color="CBD5E1")
    thick_bottom = Side(style="medium", color="0F172A")
    double_bottom = Side(style="double", color="0F172A")
    border_cell = Border(left=thin_line, right=thin_line, top=thin_line, bottom=thin_line)
    border_kpi = Border(left=thin_line, right=thin_line, top=thin_line, bottom=thin_line)
    border_header = Border(left=thin_line, right=thin_line, top=thin_line, bottom=thick_bottom)
    border_total = Border(top=thin_line, bottom=double_bottom)
    
    # Alineaciones
    al_center = Alignment(horizontal="center", vertical="center")
    al_left = Alignment(horizontal="left", vertical="center")
    al_right = Alignment(horizontal="right", vertical="center")

    # ----------------------------------------------------
    # FILTRADO DE DATOS
    # ----------------------------------------------------
    qs_med = MedicionOpacidad.objects.filter(anulado=False).select_related(
        "equipo", "equipo__tipo", "opacimetro"
    ).order_by("-fecha", "equipo__tipo__codigo", "equipo__numero")
    
    if ano:
        qs_med = qs_med.filter(fecha__year=ano)
    if periodo:
        qs_med = qs_med.filter(periodo=periodo)
    if tipo_codigo:
        qs_med = qs_med.filter(equipo__tipo__codigo=tipo_codigo)
        
    mediciones = list(qs_med)
    
    qs_proy = ProyeccionOpacidad.objects.select_related(
        "equipo", "equipo__tipo"
    ).order_by("equipo__tipo__codigo", "equipo__numero")
    if tipo_codigo:
        qs_proy = qs_proy.filter(equipo__tipo__codigo=tipo_codigo)
    proyecciones = list(qs_proy)

    total_equipos_flota = Equipo.objects.exclude(estado="fuera_servicio").count()
    if tipo_codigo:
        total_equipos_flota = Equipo.objects.filter(tipo__codigo=tipo_codigo).exclude(estado="fuera_servicio").count()

    total_tests = len(mediciones)
    tests_aprobados = sum(1 for m in mediciones if m.aprobado)
    tests_reprobados = total_tests - tests_aprobados
    cumplimiento_pct = (tests_aprobados / total_tests) if total_tests else 0.0
    
    equipos_controlados_ids = set(m.equipo_id for m in mediciones)
    cobertura_flota_pct = (len(equipos_controlados_ids) / total_equipos_flota) if total_equipos_flota else 0.0
    tests_calibracion_vigente = sum(1 for m in mediciones if m.calibracion_vigente)

    cfg = ConfiguracionSistema.load()

    # ====================================================
    # HOJA 1: RESUMEN EJECUTIVO
    # ====================================================
    ws_res = wb.active
    ws_res.title = "Resumen Ejecutivo"
    ws_res.views.sheetView[0].showGridLines = True
    
    # Encabezado Institucional
    ws_res["B2"] = "SAN ANTONIO TERMINAL INTERNACIONAL (STI) — TIREWATCH"
    ws_res["B2"].font = f_subtitle
    ws_res["B3"] = "INFORME DE AUDITORÍA Y CONTROL DE EMISIONES DIÉSEL (OPACIDAD)"
    ws_res["B3"].font = f_title
    
    # Subtítulo con filtros y fecha de emisión
    filtro_txt = []
    if ano: filtro_txt.append(f"Año: {ano}")
    if periodo: filtro_txt.append(f"Periodo: {periodo}")
    if tipo_codigo: filtro_txt.append(f"Flota: {tipo_codigo}")
    filtro_str = " · ".join(filtro_txt) if filtro_txt else "Histórico Completo de la Flota"
    
    ws_res["B4"] = f"Alcance: {filtro_str} | Emitido: {timezone.localtime().strftime('%d/%m/%Y %H:%M')} | Generado por: {usuario.get_full_name() if usuario and hasattr(usuario, 'get_full_name') and usuario.get_full_name() else 'Sistema TireWatch'}"
    ws_res["B4"].font = f_subtitle

    # --- TARJETAS KPI (Filas 6 a 8) ---
    kpis = [
        ("TOTAL MEDICIONES", f"{total_tests:,}", "Tests realizados"),
        ("CUMPLIMIENTO NORMATIVO", f"{cumplimiento_pct:.1%}", f"{tests_aprobados} de {total_tests} aprobados"),
        ("COBERTURA DE FLOTA", f"{cobertura_flota_pct:.1%}", f"{len(equipos_controlados_ids)} de {total_equipos_flota} equipos"),
        ("INSTRUMENTAL VIGENTE", f"{(tests_calibracion_vigente/total_tests):.1%}" if total_tests else "0.0%", "Calibración al día"),
        ("TESTS REPROBADOS", f"{tests_reprobados:,}", "Requieren mantención")
    ]
    
    col_start = 2
    for title, val, sub in kpis:
        c1 = col_start
        c2 = col_start + 1
        ws_res.merge_cells(start_row=6, start_column=c1, end_row=6, end_column=c2)
        ws_res.merge_cells(start_row=7, start_column=c1, end_row=7, end_column=c2)
        ws_res.merge_cells(start_row=8, start_column=c1, end_row=8, end_column=c2)
        
        ws_res.cell(row=6, column=c1, value=title).font = f_kpi_lbl
        ws_res.cell(row=7, column=c1, value=val).font = f_kpi_num
        ws_res.cell(row=8, column=c1, value=sub).font = f_subtitle
        
        for r in range(6, 9):
            for c in range(c1, c2 + 1):
                cell = ws_res.cell(row=r, column=c)
                cell.fill = fill_kpi_card
                cell.border = border_kpi
                cell.alignment = al_center
        col_start += 2

    # --- TABLA 1: DESGLOSE POR TIPO DE FLOTA (Fila 11) ---
    ws_res["B10"] = "1. Resumen de Emisiones por Tipo de Flota"
    ws_res["B10"].font = f_section
    
    headers_flota = [
        "Tipo de Equipo", "Código", "Total Flota", "Equipos Controlados",
        "Cobertura %", "Total Tests", "Aprobados", "Reprobados", "Cumplimiento %",
        "k Medio Prom.", "Límite Estándar"
    ]
    
    for c_idx, h in enumerate(headers_flota, start=2):
        cell = ws_res.cell(row=11, column=c_idx, value=h)
        cell.font = f_header
        cell.fill = fill_header_navy
        cell.alignment = al_center
        cell.border = border_header
        
    r_idx = 12
    for tipo in TipoEquipo.objects.all():
        eq_tipo = Equipo.objects.filter(tipo=tipo).exclude(estado="fuera_servicio")
        tot_eq = eq_tipo.count()
        if tot_eq == 0 and not MedicionOpacidad.objects.filter(equipo__tipo=tipo).exists():
            continue
            
        meds_tipo = [m for m in mediciones if m.equipo.tipo_id == tipo.id]
        eq_ctrl = len(set(m.equipo_id for m in meds_tipo))
        cob_pct = (eq_ctrl / tot_eq) if tot_eq else 0.0
        n_meds = len(meds_tipo)
        aprob = sum(1 for m in meds_tipo if m.aprobado)
        reprob = n_meds - aprob
        cumpl = (aprob / n_meds) if n_meds else 0.0
        k_prom = (sum(m.k_medio for m in meds_tipo) / n_meds) if n_meds else 0.0
        limite_std = cfg.get_limite_opacidad(tipo.codigo)
        
        row_data = [
            tipo.nombre, tipo.codigo, tot_eq, eq_ctrl,
            cob_pct, n_meds, aprob, reprob, cumpl,
            k_prom, limite_std
        ]
        
        for c_idx, val in enumerate(row_data, start=2):
            cell = ws_res.cell(row=r_idx, column=c_idx, value=val)
            cell.font = f_normal
            cell.border = border_cell
            
            # Formatos
            if c_idx in (2, 3):
                cell.alignment = al_left
            elif c_idx in (4, 5, 7, 8, 9):
                cell.alignment = al_center
                cell.number_format = "#,##0"
            elif c_idx in (6, 10):
                cell.alignment = al_center
                cell.number_format = "0.0%"
            elif c_idx in (11, 12):
                cell.alignment = al_center
                cell.number_format = "0.00"
        r_idx += 1

    # --- TABLA 2: DESGLOSE POR PERIODO SEMESTRAL (Fila r_idx + 2) ---
    r_idx += 2
    ws_res.cell(row=r_idx, column=2, value="2. Histórico de Cumplimiento por Campaña Semestral").font = f_section
    r_idx += 1
    
    headers_periodo = [
        "Periodo / Semestre", "Año", "Semestre", "Total Tests",
        "Aprobados", "Reprobados", "% Cumplimiento", "k Medio Flota",
        "Calibración Vigente %", "Equipos Evaluados"
    ]
    
    for c_idx, h in enumerate(headers_periodo, start=2):
        cell = ws_res.cell(row=r_idx, column=c_idx, value=h)
        cell.font = f_header
        cell.fill = fill_header_cyan
        cell.alignment = al_center
        cell.border = border_header
        
    r_idx += 1
    periodos_disponibles = sorted(list(set(m.periodo for m in mediciones if m.periodo)), reverse=True)
    for per in periodos_disponibles:
        meds_p = [m for m in mediciones if m.periodo == per]
        n_p = len(meds_p)
        ap_p = sum(1 for m in meds_p if m.aprobado)
        rep_p = n_p - ap_p
        cumpl_p = (ap_p / n_p) if n_p else 0.0
        k_prom_p = (sum(m.k_medio for m in meds_p) / n_p) if n_p else 0.0
        cal_p = (sum(1 for m in meds_p if m.calibracion_vigente) / n_p) if n_p else 0.0
        eqs_p = len(set(m.equipo_id for m in meds_p))
        
        ano_p = per.split("-")[0] if "-" in per else ""
        sem_p = per.split("-")[1] if "-" in per else ""
        
        row_data = [
            per, ano_p, sem_p, n_p,
            ap_p, rep_p, cumpl_p, k_prom_p,
            cal_p, eqs_p
        ]
        
        for c_idx, val in enumerate(row_data, start=2):
            cell = ws_res.cell(row=r_idx, column=c_idx, value=val)
            cell.font = f_normal
            cell.border = border_cell
            cell.alignment = al_center
            
            if c_idx in (5, 6, 7, 11):
                cell.number_format = "#,##0"
            elif c_idx in (8, 10):
                cell.number_format = "0.0%"
            elif c_idx == 9:
                cell.number_format = "0.00"
        r_idx += 1

    # Auto-ajuste de columnas de Hoja 1
    for col in ws_res.columns:
        col_letter = get_column_letter(col[0].column)
        if col_letter == 'A':
            ws_res.column_dimensions[col_letter].width = 4
            continue
        max_len = 0
        for cell in col:
            val_str = str(cell.value or '')
            if cell.row > 4 and len(val_str) > max_len:
                max_len = len(val_str)
        ws_res.column_dimensions[col_letter].width = max(max_len + 4, 15)

    # ====================================================
    # HOJA 2: DETALLE MEDICIONES HISTÓRICAS
    # ====================================================
    ws_det = wb.create_sheet(title="Histórico de Mediciones")
    ws_det.views.sheetView[0].showGridLines = True
    
    headers_det = [
        "ID", "Código Equipo", "Tipo Equipo", "N° Equipo", "Fecha Test",
        "Hora", "Año", "Periodo", "Horómetro (hrs)", "k Medio (1/m)",
        "k Límite (1/m)", "% del Límite", "Margen (1/m)", "Resultado",
        "Temp. Aceite (°C)", "RPM Ralentí", "RPM Máximo", "Opacímetro N° Serie",
        "Modelo Instrumento", "Calibración Vigente", "Venc. Calibración",
        "Operador / Inspector", "Archivo PDF Adjunto", "Fecha Registro"
    ]
    
    for c_idx, h in enumerate(headers_det, start=1):
        cell = ws_det.cell(row=1, column=c_idx, value=h)
        cell.font = f_header
        cell.fill = fill_header_navy
        cell.alignment = al_center
        cell.border = border_header
        
    for r_idx, m in enumerate(mediciones, start=2):
        pct = m.pct_limite if m.pct_limite is not None else (m.k_medio / m.k_limite if m.k_limite else 0.0)
        margen = m.margen if m.margen is not None else (m.k_limite - m.k_medio if m.k_limite else 0.0)
        
        row_vals = [
            m.id,
            m.equipo.codigo_completo,
            m.equipo.tipo.nombre,
            m.equipo.numero,
            m.fecha.strftime("%Y-%m-%d") if m.fecha else "",
            m.hora.strftime("%H:%M") if m.hora else "",
            m.fecha.year if m.fecha else "",
            m.periodo,
            m.horometro_motor or "—",
            m.k_medio,
            m.k_limite,
            pct,
            margen,
            "APROBADO" if m.aprobado else "REPROBADO",
            m.temp_aceite or "—",
            m.rpm_ralenti or "—",
            m.rpm_maximo or "—",
            m.opacimetro.numero_serie if m.opacimetro else "—",
            m.opacimetro.modelo if m.opacimetro else "—",
            "VIGENTE" if m.calibracion_vigente else "VENCIDA",
            m.opacimetro.vencimiento_control.strftime("%Y-%m-%d") if m.opacimetro and m.opacimetro.vencimiento_control else "—",
            m.operador or "—",
            m.archivo.name.split("/")[-1] if m.archivo else "—",
            m.created_at.strftime("%Y-%m-%d %H:%M") if m.created_at else ""
        ]
        
        fill_row = fill_zebra if r_idx % 2 == 0 else None
        
        for c_idx, val in enumerate(row_vals, start=1):
            cell = ws_det.cell(row=r_idx, column=c_idx, value=val)
            cell.font = f_normal
            cell.border = border_cell
            if fill_row: cell.fill = fill_row
            
            # Alineación y formatos
            if c_idx in (1, 4, 7, 8):
                cell.alignment = al_center
            elif c_idx in (2, 3, 18, 19, 22, 23):
                cell.alignment = al_left
            elif c_idx in (5, 6, 20, 21, 24):
                cell.alignment = al_center
            elif c_idx in (9, 15, 16, 17):
                cell.alignment = al_right
                if isinstance(val, (int, float)): cell.number_format = "#,##0"
            elif c_idx in (10, 11, 13):
                cell.alignment = al_right
                cell.number_format = "0.00"
            elif c_idx == 12:
                cell.alignment = al_right
                cell.number_format = "0.0%"
                
            # Destacar Aprobado / Reprobado
            if c_idx == 14:
                cell.alignment = al_center
                if val == "APROBADO":
                    cell.fill = fill_green
                    cell.font = f_green
                else:
                    cell.fill = fill_red
                    cell.font = f_red
                    
            # Destacar Calibración
            if c_idx == 20:
                if val == "VIGENTE":
                    cell.font = f_green
                else:
                    cell.font = f_red
                    cell.fill = fill_yellow

    for col in ws_det.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws_det.column_dimensions[col_letter].width = max(max_len + 3, 11)

    # ====================================================
    # HOJA 3: ESTADO ACTUAL DE LA FLOTA
    # ====================================================
    ws_flota = wb.create_sheet(title="Estado Actual de la Flota")
    ws_flota.views.sheetView[0].showGridLines = True
    
    headers_flota_st = [
        "Código Equipo", "Tipo", "N° Equipo", "Estado Predictivo",
        "k Actual (1/m)", "k Límite (1/m)", "% del Límite", "Tasa Anual (1/m/año)",
        "k Proyectado (+6m)", "N° Mediciones", "Último Control", "Próximo Control",
        "Estado Control", "Diagnóstico / Motivo del Semáforo"
    ]
    
    for c_idx, h in enumerate(headers_flota_st, start=1):
        cell = ws_flota.cell(row=1, column=c_idx, value=h)
        cell.font = f_header
        cell.fill = fill_header_navy
        cell.alignment = al_center
        cell.border = border_header
        
    for r_idx, p in enumerate(proyecciones, start=2):
        row_vals = [
            p.equipo.codigo_completo,
            p.equipo.tipo.nombre,
            p.equipo.numero,
            p.estado,
            p.k_actual if p.k_actual is not None else "—",
            p.k_limite if p.k_limite is not None else "—",
            p.pct_limite if p.pct_limite is not None else "—",
            p.tasa_k_anual if p.tasa_k_anual is not None else "—",
            p.k_proyectado if p.k_proyectado is not None else "—",
            p.n_mediciones,
            p.ultima_medicion_fecha.strftime("%Y-%m-%d") if p.ultima_medicion_fecha else "—",
            p.proximo_control.strftime("%Y-%m-%d") if p.proximo_control else "—",
            "VENCIDO" if p.control_vencido else "AL DÍA",
            p.motivo_estado or "—"
        ]
        
        fill_row = fill_zebra if r_idx % 2 == 0 else None
        
        for c_idx, val in enumerate(row_vals, start=1):
            cell = ws_flota.cell(row=r_idx, column=c_idx, value=val)
            cell.font = f_normal
            cell.border = border_cell
            if fill_row: cell.fill = fill_row
            
            if c_idx in (1, 2, 14):
                cell.alignment = al_left
            elif c_idx in (3, 10, 11, 12):
                cell.alignment = al_center
            elif c_idx in (5, 6, 8, 9):
                cell.alignment = al_right
                if isinstance(val, (int, float)): cell.number_format = "0.00"
            elif c_idx == 7:
                cell.alignment = al_right
                if isinstance(val, (int, float)): cell.number_format = "0.0%"
                
            # Semáforo Estado
            if c_idx == 4:
                cell.alignment = al_center
                if val == "OK":
                    cell.fill = fill_green
                    cell.font = f_green
                elif val == "ATENCION":
                    cell.fill = fill_yellow
                    cell.font = f_yellow
                elif val == "CRITICO":
                    cell.fill = fill_red
                    cell.font = f_red
                    
            # Estado Control
            if c_idx == 13:
                cell.alignment = al_center
                if val == "AL DÍA":
                    cell.font = f_green
                else:
                    cell.font = f_red
                    cell.fill = fill_yellow

    for col in ws_flota.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws_flota.column_dimensions[col_letter].width = max(max_len + 3, 12)

    # Guardar en memoria
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
